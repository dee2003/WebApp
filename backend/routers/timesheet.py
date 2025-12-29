from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session, joinedload
from typing import List,Optional
from .. import models, schemas
from ..database import get_db
import os
from .. import models, schemas # This line already imports your schemas
import pandas as pd
import json
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy.orm.attributes import flag_modified
import os
from ..models import Timesheet,JobPhase   # <-- import your model
from ..database import get_db
from ..auditing import log_action
from .. import oauth2

import httpx

load_dotenv()
load_dotenv(r"D:\WebApp\backend\.env")  # Use raw string for Windows
sender_email = os.getenv("SMTP_USER")
sender_password = os.getenv("SMTP_PASSWORD")
print("SMTP_USER:", sender_email)
print("SMTP_PASSWORD:", sender_password)

router = APIRouter(
    prefix="/api/timesheets",
    tags=["Timesheets"]
)
from datetime import datetime
from sqlalchemy import func, case
from .. import models, schemas, database
from sqlalchemy import String, literal


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(BASE_DIR, ".env")
load_dotenv(dotenv_path=ENV_PATH)
BASE_URL = os.getenv("BASE_URL")
print("🔗 [Router] BASE_URL loaded:", BASE_URL)

MAPBOX_TOKEN = os.getenv("MAPBOX_TOKEN")
OFFICE_COORDS = {"lat": 38.9072, "lon": -77.0369}

async def get_mapbox_distance(address: str):
    """Securely calculates distance on the server side."""
    if not address or address == "No Address":
        return 0
    try:
        async with httpx.AsyncClient() as client:
            geo_url = f"https://api.mapbox.com/geocoding/v5/mapbox.places/{address}.json?access_token={MAPBOX_TOKEN}"
            geo_res = await client.get(geo_url)
            features = geo_res.json().get("features", [])
            if not features: return 0
            
            dest_lon, dest_lat = features[0]["center"]

            dir_url = f"https://api.mapbox.com/directions/v5/mapbox/driving/{OFFICE_COORDS['lon']},{OFFICE_COORDS['lat']};{dest_lon},{dest_lat}?access_token={MAPBOX_TOKEN}"
            dir_res = await client.get(dir_url)
            routes = dir_res.json().get("routes", [])
            return routes[0]["distance"] if routes else 0
    except Exception:
        return 0

@router.get("/morning-brief")
async def get_morning_brief(db: Session = Depends(get_db)):
    """Enriched endpoint for Natalia's Dashboard."""
    timesheets = db.query(models.Timesheet).filter(
        models.Timesheet.status == "DRAFT"
    ).options(joinedload(models.Timesheet.foreman)).all()
    
    results = []
    for ts in timesheets:
        f_name = f"{ts.foreman.first_name} {ts.foreman.last_name}" if ts.foreman else "Unknown"
        
        # Build nested data Category -> Vendor -> Materials
        categorized_data = {"Concrete": {}, "Asphalt": {}, "Top Soil": {}, "Trucking": {}}
        
        # 1. Process Vendors
        vendor_mats = ts.data.get("selected_vendor_materials", {})
        for v_id, v_data in vendor_mats.items():
            v_name = v_data.get("name", "Vendor")
            cat = v_data.get("vendor_category", "General")
            if "Asphalt" in cat: cat = "Asphalt"
            
            if cat in categorized_data:
                if v_name not in categorized_data[cat]: categorized_data[cat][v_name] = []
                for m in v_data.get("selectedMaterials", []):
                    if m.get('detail'):
                        categorized_data[cat][v_name].append({
                            "important": m.get('detail').strip(),
                            "material": m.get('material')
                        })

        # 2. Process Trucking
        trucking = ts.data.get("selected_material_items", {})
        for t_id, t_data in trucking.items():
            t_name = t_data.get("name", "Hauler")
            if t_data.get("notes"):
                if t_name not in categorized_data["Trucking"]: categorized_data["Trucking"][t_name] = []
                categorized_data["Trucking"][t_name].append({
                    "important": t_data.get("notes").strip(),
                    "material": "Trucking Services"
                })

        # Distance logic (Await async helper)
        address = ts.data.get("location", "No Address")
        dist_meters = await get_mapbox_distance(address)

        results.append({
            "id": ts.id,
            "status": ts.status,
            "date": str(ts.date),
            "brief": {
                "foreman": f_name,
                "job_code": ts.data.get("job", {}).get("job_code", "N/A"),
                "job_name": ts.data.get("job_name", "N/A"),
                "categorized_data": categorized_data,
                "address": address,
                "distanceMeters": dist_meters,
                "distanceMiles": round(dist_meters / 1609.34, 1) if dist_meters > 0 else "N/A"
            }
        })
    return results

    
# def format_natalia_alert(ts_data: dict, foreman_name: str):
#     # Initialize nested structure
#     categorized_data = {
#         "Concrete": {},
#         "Asphalt": {},
#         "Top Soil": {},
#         "Trucking": {}
#     }
    
#     # 1. Process Vendors
#     vendor_materials = ts_data.get("selected_vendor_materials", {})
#     for v_id, v_data in vendor_materials.items():
#         v_name = v_data.get("name", "Vendor")
#         cat_key = v_data.get("vendor_category", "General")
#         # Normalize category keys
#         if "Asphalt" in cat_key: cat_key = "Asphalt"
        
#         if cat_key in categorized_data:
#             if v_name not in categorized_data[cat_key]:
#                 categorized_data[cat_key][v_name] = []
            
#             for m in v_data.get("selectedMaterials", []):
#                 if m.get('detail'):
#                     categorized_data[cat_key][v_name].append({
#                         "important": m.get('detail').strip(),
#                         "material": m.get('material')
#                     })

#     # 2. Process Trucking
#     trucking_items = ts_data.get("selected_material_items", {})
#     for t_id, t_data in trucking_items.items():
#         t_name = t_data.get("name", "Hauler")
#         notes = t_data.get("notes", "")
#         if notes:
#             if t_name not in categorized_data["Trucking"]:
#                 categorized_data["Trucking"][t_name] = []
#             categorized_data["Trucking"][t_name].append({
#                 "important": notes.strip(),
#                 "material": "Trucking"
#             })

#     return {
#         "foreman": foreman_name,
#         "job_code": ts_data.get("job", {}).get("job_code", "N/A"),
#         "job_name": ts_data.get("job_name", "N/A"),
#         "categorized_data": categorized_data,
#         "address": ts_data.get("location", "No Address")
#     }

# @router.get("/morning-brief")
# def get_morning_brief(db: Session = Depends(get_db)):
#     # Optimized to include status and exact date for Natalia's Dashboard
#     timesheets = db.query(models.Timesheet).filter(
#         models.Timesheet.status == "DRAFT" # Dispatcher's scheduled drafts
#     ).options(joinedload(models.Timesheet.foreman)).all()
    
#     briefs = []
#     for ts in timesheets:
#         f_name = f"{ts.foreman.first_name} {ts.foreman.last_name}" if ts.foreman else "Unknown"
#         briefs.append({
#             "id": ts.id,
#             "status": ts.status,
#             "date": ts.date.strftime("%Y-%m-%d") if hasattr(ts.date, "strftime") else str(ts.date),
#             "brief": format_natalia_alert(ts.data, f_name)
#         })
#     return briefs



@router.get("/counts-by-status", response_model=schemas.TimesheetCountsResponse)
def get_timesheet_counts_by_status(db: Session = Depends(get_db)):
    try:
        foreman_statuses = [
            models.SubmissionStatus.DRAFT.value,
            models.SubmissionStatus.PENDING.value,
        ]
        supervisor_statuses = [
            models.SubmissionStatus.SUBMITTED.value,
        ]

        engineer_status = models.SubmissionStatus.APPROVED_BY_SUPERVISOR.value

        counts_query = db.query(
            func.count(
                case((models.Timesheet.status.cast(String).in_(foreman_statuses), 1))
            ).label("foreman_total"),
            func.count(
                case((models.Timesheet.status.cast(String).in_(supervisor_statuses), 1))
            ).label("supervisor_total"),
            func.count(
                case((models.Timesheet.status.cast(String) == engineer_status, 1))
            ).label("engineer_total"),
        ).first()
        return {
            "foreman": int(counts_query.foreman_total or 0),
            "supervisor": int(counts_query.supervisor_total or 0),
            "project_engineer": int(counts_query.engineer_total or 0),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

import httpx 
MAPBOX_TOKEN = os.getenv("MAPBOX_TOKEN")
OFFICE_COORDS = {"lat": 38.9072, "lon": -77.0369}
async def get_mapbox_distance(address: str):
    """Securely calculates distance on the server side."""
    if not address or address == "No Address":
        return 0
    try:
        async with httpx.AsyncClient() as client:
            geo_url = f"https://api.mapbox.com/geocoding/v5/mapbox.places/{address}.json?access_token={MAPBOX_TOKEN}"
            geo_res = await client.get(geo_url)
            features = geo_res.json().get("features", [])
            if not features: return 0
            dest_lon, dest_lat = features[0]["center"]
            dir_url = f"https://api.mapbox.com/directions/v5/mapbox/driving/{OFFICE_COORDS['lon']},{OFFICE_COORDS['lat']};{dest_lon},{dest_lat}?access_token={MAPBOX_TOKEN}"
            dir_res = await client.get(dir_url)
            routes = dir_res.json().get("routes", [])
            return routes[0]["distance"] if routes else 0
    except Exception:
        return 0
@router.get("/morning-brief")
async def get_morning_brief(db: Session = Depends(get_db)):
    """Enriched endpoint for Natalia's Dashboard."""
    timesheets = db.query(models.Timesheet).filter(
        models.Timesheet.status == "DRAFT"
    ).options(joinedload(models.Timesheet.foreman)).all()
    results = []
    for ts in timesheets:
        f_name = f"{ts.foreman.first_name} {ts.foreman.last_name}" if ts.foreman else "Unknown"
        # Build nested data Category -> Vendor -> Materials
        categorized_data = {"Concrete": {}, "Asphalt": {}, "Top Soil": {}, "Trucking": {}}
        # 1. Process Vendors
        vendor_mats = ts.data.get("selected_vendor_materials", {})
        for v_id, v_data in vendor_mats.items():
            v_name = v_data.get("name", "Vendor")
            cat = v_data.get("vendor_category", "General")
            if "Asphalt" in cat: cat = "Asphalt"
            if cat in categorized_data:
                if v_name not in categorized_data[cat]: categorized_data[cat][v_name] = []
                for m in v_data.get("selectedMaterials", []):
                    if m.get('detail'):
                        categorized_data[cat][v_name].append({
                            "important": m.get('detail').strip(),
                            "material": m.get('material')
                        })
        # 2. Process Trucking
        trucking = ts.data.get("selected_material_items", {})
        for t_id, t_data in trucking.items():
            t_name = t_data.get("name", "Hauler")
            if t_data.get("notes"):
                if t_name not in categorized_data["Trucking"]: categorized_data["Trucking"][t_name] = []
                categorized_data["Trucking"][t_name].append({
                    "important": t_data.get("notes").strip(),
                    "material": "Trucking Services"
                })
        # Distance logic (Await async helper)
        address = ts.data.get("location", "No Address")
        dist_meters = await get_mapbox_distance(address)
        results.append({
            "id": ts.id,
            "status": ts.status,
            "date": str(ts.date),
            "brief": {
                "foreman": f_name,
                "job_code": ts.data.get("job", {}).get("job_code", "N/A"),
                "job_name": ts.data.get("job_name", "N/A"),
                "categorized_data": categorized_data,
                "address": address,
                "distanceMeters": dist_meters,
                "distanceMiles": round(dist_meters / 1609.34, 1) if dist_meters > 0 else "N/A"
            }
        })
    return results
@router.post("/", response_model=schemas.Timesheet)

def create_timesheet(
    timesheet: schemas.TimesheetCreate,
    db: Session = Depends(get_db)
):
    print("FULL PAYLOAD:", timesheet.model_dump())
    print("job_phase_id:", timesheet.job_phase_id)

    data_to_store = timesheet.data or {}

    # --- Derive job name robustly ---
    job_name = (
        data_to_store.get("job_name")
        or (data_to_store.get("job") or {}).get("job_description")
        or (data_to_store.get("job") or {}).get("job_name")
        or (data_to_store.get("job") or {}).get("job_code")
        or "Untitled Timesheet"
    )

    # --- Create new timesheet ---
    db_ts = models.Timesheet(
        foreman_id=timesheet.foreman_id,
        job_phase_id=timesheet.job_phase_id,
        date=timesheet.date,
        status="DRAFT",
        data=data_to_store,
        timesheet_name=job_name
    )

    db.add(db_ts)
    db.commit()
    db.refresh(db_ts)

    # --- AUDIT LOG ---
# Fetch the related JobPhase
    job_phase = db.query(JobPhase).filter(JobPhase.id == timesheet.job_phase_id).first()
    job_code = job_phase.job_code if job_phase else "Unknown"

    # Audit log
    log_action(
        db=db,
        user_id=timesheet.foreman_id,
        action="CREATED",
        target_resource="TIMESHEET",
        target_resource_id=str(db_ts.id),
        details=(
            f"Timesheet '{db_ts.timesheet_name}' "
            f"created for job_code={job_code} "
            f"on date={timesheet.date}"
        )
    )


    db.commit()
    return db_ts


@router.get("/by-foreman/{foreman_id}", response_model=List[schemas.Timesheet])
def get_timesheets_by_foreman(foreman_id: int, db: Session = Depends(get_db)):
    """
    Returns only editable timesheets (Draft or Pending) for a given foreman.
    'Sent' or 'Approved' timesheets will no longer appear in the app list.
    """
    timesheets = (
        db.query(models.Timesheet)
        .options(joinedload(models.Timesheet.files))
        .filter(
            models.Timesheet.foreman_id == foreman_id,
            models.Timesheet.status.in_([
                models.SubmissionStatus.DRAFT,
                models.SubmissionStatus.PENDING
            ])
        )
        .order_by(models.Timesheet.date.desc())
        .all()
    )
    return timesheets


from datetime import date as date_type
from sqlalchemy import cast, Date
# @router.get("/for-supervisor", response_model=List[schemas.Timesheet])
# def get_timesheets_for_supervisor(
#     db: Session = Depends(get_db),
#     foreman_id: Optional[int] = Query(None),
#     date: Optional[str] = Query(None),
# ):
#     """
#     Returns all SUBMITTED timesheets for supervisors to review.
#     Filters by foreman_id and/or work date (timesheet.date).
#     """
#     # Use "status" instead of "sent"
#     query = (
#         db.query(models.Timesheet)
#         .options(joinedload(models.Timesheet.files))
#         .filter(models.Timesheet.status == "SUBMITTED")
#     )

#     if foreman_id is not None:
#         query = query.filter(models.Timesheet.foreman_id == foreman_id)

#     if date:
#         try:
#             target_date = date_type.fromisoformat(date)
#             query = query.filter(cast(models.Timesheet.date, Date) == target_date)
#         except ValueError:
#             raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

#     # Optional: latest submitted first
#     timesheets = query.order_by(models.Timesheet.updated_at.desc()).all()

#     return timesheets
@router.get("/for-supervisor", response_model=List[schemas.Timesheet])
def get_timesheets_for_supervisor(
    db: Session = Depends(get_db),
    foreman_id: Optional[int] = Query(None),
    date: Optional[str] = Query(None),
):
    # Change the filter to allow both SUBMITTED and REVIEWED statuses
    query = (
        db.query(models.Timesheet)
        .options(joinedload(models.Timesheet.files))
        .filter(
            models.Timesheet.status.in_([
                "SUBMITTED", 
                "REVIEWED_BY_SUPERVISOR"
            ])
        )
    )

    if foreman_id is not None:
        query = query.filter(models.Timesheet.foreman_id == foreman_id)

    if date:
        try:
            target_date = date_type.fromisoformat(date)
            query = query.filter(cast(models.Timesheet.date, Date) == target_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format.")

    return query.order_by(models.Timesheet.updated_at.desc()).all()


from ..schemas import TimesheetWorkflowSchema
from ..models import TimesheetWorkflow
@router.get("/workflows", response_model=List[TimesheetWorkflowSchema])
def get_timesheet_workflows(timesheet_id: int, db: Session = Depends(get_db)):
    return (
        db.query(TimesheetWorkflow)
        .filter(TimesheetWorkflow.timesheet_id == timesheet_id)
        .order_by(TimesheetWorkflow.timestamp.desc())
        .all()
    )

@router.get("/{timesheet_id}")
def get_single_timesheet(timesheet_id: int, db: Session = Depends(get_db)):
    """
    Returns a single timesheet, prioritizing saved JSON data and only enriching
    non-conflicting fields from static database models.
    """
    timesheet = (
        db.query(models.Timesheet)
        .options(joinedload(models.Timesheet.files))
        .filter(models.Timesheet.id == timesheet_id)
        .first()
    )
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    saved_data = timesheet.data or {}
    if isinstance(saved_data, str):
        try:
            saved_data = json.loads(saved_data)
        except json.JSONDecodeError:
            saved_data = {}
    # --- ENRICHMENT HELPER FUNCTION ---
    def enrich_entities(entity_key: str, model, name_fields: list, add_phase_defaults: bool = False, skip_name_enrichment: bool = False):
        source_data = saved_data.get(entity_key, [])
        if not source_data:
            return
        # 1. Collect all entity IDs to query the static database table once.
        entity_ids = {str(e.get("id")) for e in source_data if isinstance(e, dict) and e.get("id")}
        if not entity_ids:
            return
        db_entities = db.query(model).filter(model.id.in_(entity_ids)).all()
        db_map = {str(db_e.id): db_e for db_e in db_entities}
        phase_codes = saved_data.get("job", {}).get("phase_codes", [])
        enriched_list = []
        for entity in source_data:
            # 2. CRITICAL: Create a shallow copy to retain ALL saved keys
            # (including hours_per_phase, tickets_per_phase, and the saved 'name').
            enriched_entity = entity.copy()
            eid = str(enriched_entity.get("id"))
            db_record = db_map.get(eid)
            # 3. Preserve existing phase fields or add defaults if missing
            if add_phase_defaults:
                # Merge saved hours with DB phases
                existing_hours = enriched_entity.get("hours_per_phase", {})
                merged_hours = {p: existing_hours.get(p, 0) for p in phase_codes}
                enriched_entity["hours_per_phase"] = merged_hours

            # Preserve tickets_loads if present
            if "tickets_loads" in enriched_entity:
                enriched_entity["tickets_loads"] = enriched_entity.get("tickets_loads", {})
            

            # 4. Apply Database Enrichment
            if db_record:
                if not skip_name_enrichment:
                    # Logic for Employees (FirstName/LastName) or Equipment (Name)
                    # Compose and update 'name' field from DB attributes
                    name_parts = [getattr(db_record, f, "") for f in name_fields]
                    enriched_entity["name"] = " ".join(filter(None, name_parts)).strip()
                    # Update other non-name fields (like 'status' for employee)
                    for field in name_fields:
                        if field not in ("first_name", "last_name", "name"):
                            enriched_entity[field] = getattr(db_record, field, enriched_entity.get(field))
                # If skip_name_enrichment is TRUE (for vendors/dumping sites),
                # the original saved name and all phase data are automatically preserved
                # by the initial 'entity.copy()' in step 2.
            enriched_list.append(enriched_entity)
        saved_data[entity_key] = enriched_list
    # --- EXECUTE ENRICHMENT ---
    try:
        if hasattr(models, "Employee"):
            # Employee names must be rebuilt from DB fields (First Name, Last Name)
            enrich_entities("employees", models.Employee, ["first_name", "last_name", "status"])
        if hasattr(models, "Equipment"):
            # Equipment names are usually enriched/validated
            enrich_entities("equipment", models.Equipment, ["name"])
        # VENDORS: skip_name_enrichment=True ensures saved name and phase data are kept.
        if hasattr(models, "Vendor"):
            enrich_entities("vendors", models.Vendor, ["name"], add_phase_defaults=False)
            
        # MATERIALS_TRUCKING: Assuming standard enrichment (no skip).
        if hasattr(models, "MaterialTrucking"):
            enrich_entities("materials_trucking", models.MaterialTrucking, ["name"], add_phase_defaults=False)
            
        # DUMPING_SITES: skip_name_enrichment=True ensures saved name and phase data are kept.
        if hasattr(models, "DumpingSite"):
            enrich_entities("dumping_sites", models.DumpingSite, ["name"], add_phase_defaults=False)
            
    except Exception as e:
        print(f":warning: Critical enrichment error: {e}")
        # Depending on severity, you might want to re-raise the exception or log it.
    # --- FINAL RETURN ---
    return {
        "id": timesheet.id,
        "foreman_id": timesheet.foreman_id,
        "job_phase_id": timesheet.job_phase_id,
        "date": timesheet.date,
        "status": timesheet.status,
        "timesheet_name": timesheet.timesheet_name,
        "data": saved_data,
    }
from sqlalchemy.orm.attributes import flag_modified # Needed for JSONB updates
@router.put("/{timesheet_id}", response_model=schemas.Timesheet)

@router.put("/{timesheet_id}", response_model=schemas.Timesheet)
def update_timesheet(
    timesheet_id: int,
    timesheet_update: schemas.TimesheetUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(oauth2.get_current_user)
):
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    if not ts:
        raise HTTPException(status_code=404, detail="Timesheet not found")

    payload = timesheet_update.dict(exclude_unset=True)
    current_status = ts.status 

    # 1. Update Data (Always allowed for authorized users)
    if "data" in payload:
        def clean_legacy_tickets(data):
            for entity_key in ["materials_trucking", "vendors", "dumping_sites"]:
                if entity_key in data:
                    for entity in data[entity_key]:
                        entity.pop("tickets_per_phase", None)
            return data
        
        ts.data = clean_legacy_tickets(payload["data"])
        flag_modified(ts, "data") #

    # 2. Status Guardrail Logic
    # Define hierarchy order: higher index = more advanced status
    STATUS_HIERARCHY = [
        SubmissionStatus.DRAFT.value,
        SubmissionStatus.PENDING.value,
        SubmissionStatus.SUBMITTED.value,
        SubmissionStatus.REVIEWED_BY_SUPERVISOR.value,
        SubmissionStatus.APPROVED_BY_SUPERVISOR.value,
        "APPROVED_BY_PE"
    ]

    if "status" in payload:
        new_status = payload["status"]
        
        try:
            current_idx = STATUS_HIERARCHY.index(current_status)
            new_idx = STATUS_HIERARCHY.index(new_status)
            
            # ALLOW if the status is moving forward or staying the same
            if new_idx >= current_idx:
                ts.status = new_status
            else:
                # BLOCK reversion (e.g., PE tries to save, but front-end sends 'SUBMITTED')
                print(f"⚠️ [GUARDRAIL] Blocked reversion from {current_status} to {new_status}")
                # We do not update ts.status, effectively keeping the current high-level status
        except ValueError:
            # If a status is not in our hierarchy, fall back to safe assignment or ignore
            ts.status = new_status 

    # ... sync job name, commit, and excel generation logic ...
    db.commit()
    db.refresh(ts)
    # --- Keep job name synced ---
    data_to_store = ts.data or {}
    job_name = (
        data_to_store.get("job_name")
        or (data_to_store.get("job") or {}).get("job_description")
        or (data_to_store.get("job") or {}).get("job_name")
        or (data_to_store.get("job") or {}).get("job_code")
        or "Untitled Timesheet"
    )
    ts.timesheet_name = job_name
    ts.data["job_name"] = job_name
    db.commit()
    db.refresh(ts)
    log_action(
    db=db,
    user_id=current_user.id,
    action="TIMESHEET_UPDATED",
    target_resource="TIMESHEET",
    target_resource_id=str(ts.id),
    details=(
        f"Timesheet '{ts.timesheet_name}' "
        f"updated by {current_user.role.value} "
        f"({current_user.username})."
    )
)

    db.commit()

    # --- Excel file generation ---
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        storage_dir = os.path.join(BASE_DIR, "storage")
        os.makedirs(storage_dir, exist_ok=True)
        ts_date_str = ts.date.strftime("%Y-%m-%d") if hasattr(ts.date, "strftime") else str(ts.date)
        date_folder = os.path.join(storage_dir, ts_date_str)
        os.makedirs(date_folder, exist_ok=True)
        version = len([f for f in os.listdir(date_folder) if f.startswith(f"timesheet_{ts.id}_")]) + 1
        file_name = f"timesheet_{ts.id}_{ts_date_str}_v{version}.xlsx"
        file_path_local = os.path.join(date_folder, file_name)
        data = ts.data if isinstance(ts.data, dict) else json.loads(ts.data)
        job_phases = data.get("job", {}).get("phase_codes", [])
        def create_df(entities, name_key="name"):
            rows = []
            for ent in entities:
                name = ent.get(name_key) or ent.get("first_name", "")
                if "last_name" in ent:
                    name = f"{name} {ent.get('last_name', '')}".strip()
                row = {"ID": ent.get("id", ""), "Name": name}
                for phase in job_phases:
                    row[phase] = ent.get("hours_per_phase", {}).get(phase, 0)
                rows.append(row)
            return pd.DataFrame(rows)
        def create_dumping_site_df(entities):
            rows = []
            for ent in entities:
                row = {"ID": ent.get("id", ""), "Name": ent.get("name", "")}
                for phase in job_phases:
                    row[f"{phase} (Qty)"] = ent.get("hours_per_phase", {}).get(phase, 0)
                    row[f"{phase} (# of Loads)"] = ent.get("tickets_loads", {}).get(ent["id"], 0)
                rows.append(row)
            return pd.DataFrame(rows)
        with pd.ExcelWriter(file_path_local, engine="openpyxl") as writer:
            create_df(data.get("employees", []), name_key="first_name").to_excel(writer, index=False, sheet_name="Employees")
            create_df(data.get("equipment", [])).to_excel(writer, index=False, sheet_name="Equipment")
            create_df(data.get("materials_trucking", [])).to_excel(writer, index=False, sheet_name="Materials")
            create_df(data.get("vendors", [])).to_excel(writer, index=False, sheet_name="Vendors")
            create_dumping_site_df(data.get("dumping_sites", [])).to_excel(writer, index=False, sheet_name="DumpingSites")
        # :white_check_mark: Replace ngrok link with your own base URL
        # NGROK_BASE_URL = "https://coated-nonattributive-babara.ngrok-free.dev"
        # file_url = f"{NGROK_BASE_URL}/storage/{ts_date_str}/{file_name}"
        # BASE_URL = os.getenv("BASE_URL")
        NGROK_BASE_URL = "https://coated-nonattributive-babara.ngrok-free.dev"
        file_url = f"{NGROK_BASE_URL}/storage/{ts_date_str}/{file_name}"
        # :white_check_mark: Save file info in DB
        file_record = models.TimesheetFile(
            timesheet_id=ts.id,
            foreman_id=ts.foreman_id,
            file_path=file_url
        )
        db.add(file_record)
        db.commit()
        db.refresh(file_record)
        print(f":white_check_mark: Timesheet Excel generated and saved at: {file_url}")
    except Exception as e:
        print(f":x: Excel generation failed: {e}")
    return ts
# -------------------------------
# SEND a timesheet
# -------------------------------


from ..models import SubmissionStatus

from datetime import datetime
from sqlalchemy import func

@router.post("/{timesheet_id}/send", response_model=schemas.Timesheet)
def send_timesheet(
    timesheet_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(oauth2.get_current_user)
):
    # 1. Fetch timesheet
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    if not ts:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    # Trigger Natalia's Alert
    try:
        foreman_user = db.query(models.User).filter(models.User.id == ts.foreman_id).first()
        f_name = foreman_user.first_name if foreman_user else "Foreman"
        
        alert_msg = format_natalia_alert(ts.data, f_name)
        
        # Use your existing notification schema and logic
        notif = schemas.NotificationRequest(
            email="manisha.rk04@gmail.com",
            subject=f"📍 Site Alert: {ts.timesheet_name}",
            message=alert_msg
        )
        send_notification(notif)
    except Exception as e:
        print(f"Failed to notify Natalia: {e}")

    if ts.status == SubmissionStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="Timesheet already sent")

    try:
        # 2. Process Linked Tickets from the 'data' JSON field
        # We extract all ticket IDs that the foreman linked in the UI
        linked_tickets_map = ts.data.get("linked_tickets", {})
        
        all_ticket_ids = []
        if isinstance(linked_tickets_map, dict):
            for row_id, ticket_ids in linked_tickets_map.items():
                if isinstance(ticket_ids, list):
                    all_ticket_ids.extend(ticket_ids)

        # 3. Mark Tickets as SUBMITTED and link to this Timesheet ID
        if all_ticket_ids:
            # Remove duplicates just in case
            unique_ids = list(set(all_ticket_ids))
            
            db.query(models.Ticket)\
                .filter(models.Ticket.id.in_(unique_ids))\
                .update(
                    {
                        "status": SubmissionStatus.SUBMITTED, 
                        "timesheet_id": ts.id
                    }, 
                    synchronize_session=False
                )

        # 4. Update Timesheet Status
        ts.sent = True
        ts.sent_date = datetime.utcnow()
        ts.status = SubmissionStatus.SUBMITTED

        # 5. Workflow Logging
        workflow = models.TimesheetWorkflow(
            timesheet_id=ts.id,
            foreman_id=ts.foreman_id,
            action="Sent",
            by_role="Foreman",
            timestamp=datetime.utcnow(),
            comments=f"Sent to supervisor with {len(all_ticket_ids)} linked tickets",
        )
        db.add(workflow)

        # 6. Audit Log
        log_action(
            db=db,
            user_id=current_user.id,
            action="TIMESHEET_SENT",
            target_resource="TIMESHEET",
            target_resource_id=str(ts.id),
            details=f"Timesheet {ts.id} sent. {len(all_ticket_ids)} tickets updated to SUBMITTED."
        )

        db.commit()
        db.refresh(ts)
        return ts

    except Exception as e:
        db.rollback()
        print(f"Submission Error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error processing ticket submission")
# -------------------------------
# DELETE a timesheet
# -------------------------------
@router.delete("/{timesheet_id}", status_code=status.HTTP_204_NO_CONTENT)
# @audit(action="Deleted", entity="Timesheets")

def delete_timesheet(timesheet_id: int, db: Session = Depends(get_db)):
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    if not ts:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    db.delete(ts)
    db.commit()
    return
# In your routers/timesheet.py


@router.get("/", response_model=List[schemas.TimesheetResponse])
def list_timesheets(db: Session = Depends(get_db)):
    """
    Returns a list of all timesheets with foreman names and job names included.
    This is optimized for the admin dashboard view.
    """
    timesheets = db.query(models.Timesheet).options(joinedload(models.Timesheet.foreman)).all()
    
    response = []
    for ts in timesheets:
        foreman_name = f"{ts.foreman.first_name} {ts.foreman.last_name}" if ts.foreman else "N/A"
        
        # Create the response object, ensuring all required fields are present
        response.append(schemas.TimesheetResponse(
            id=ts.id,
            date=ts.date,
            foreman_id=ts.foreman_id,
            foreman_name=foreman_name,
            job_name=ts.timesheet_name,  # <-- The FIX: Populate the required 'job_name' field
            data=ts.data,
            status=ts.status
        ))
        
    return response
from sqlalchemy import or_
from ..models import SubmissionStatus  # ✅ use your enum safely

@router.get("/drafts/by-foreman/{foreman_id}", response_model=List[schemas.Timesheet])
def get_draft_timesheets_by_foreman(foreman_id: int, db: Session = Depends(get_db)):
    """
    Returns all draft/pending timesheets for a given foreman.
    This is used by the new ReviewTimesheetScreen.
    """
    timesheets = (
        db.query(models.Timesheet)
        .options(joinedload(models.Timesheet.files))
        .filter(models.Timesheet.foreman_id == foreman_id)
        .filter(
            or_(
                models.Timesheet.status == SubmissionStatus.PENDING

            )
        )
        .order_by(models.Timesheet.date.desc())
        .all()
    )
    return timesheets

@router.post("/timesheets/save-draft/")
def save_draft(timesheet: schemas.TimesheetCreate, db: Session = Depends(get_db)):
    # ... (Check for existing_ts)
    if timesheet.id:
        existing_ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet.id).first()
        if existing_ts:
            existing_ts.data = timesheet.data  # dictionary, not JSON string
            flag_modified(existing_ts, "data")  # important for JSONB change detection
            existing_ts.job_name = timesheet.job_name
            existing_ts.status = SubmissionStatus.DRAFT
            existing_ts.date = datetime.utcnow()

            db.commit()
            db.refresh(existing_ts)

            return {"message": "Draft updated", "timesheet_id": existing_ts.id}

    # Create new draft
    new_ts = models.Timesheet(
        foreman_id=timesheet.foreman_id,
        job_name=timesheet.job_name,
        date=datetime.utcnow(),
        status=SubmissionStatus.DRAFT,
        data=timesheet.data  # store dict directly
    )

    db.add(new_ts)
    db.commit()
    db.refresh(new_ts)

    return {"message": "Draft created", "timesheet_id": new_ts.id}


@router.get("/timesheets/drafts/by-foreman/{foreman_id}")
def get_drafts(foreman_id: int, db: Session = Depends(get_db)):
    drafts = (
        db.query(models.Timesheet)
        .filter(
            models.Timesheet.foreman_id == foreman_id,
            models.Timesheet.status == SubmissionStatus.DRAFT
        )
        .order_by(models.Timesheet.date.desc())
        .all()
    )

    # Convert JSONB safely (should already be dict)
    return [
        {
            "id": t.id,
            "job_name": t.job_name,
            "date": t.date,
            **(t.data if isinstance(t.data, dict) else json.loads(t.data))
        }
        for t in drafts
    ]

from typing import List, Optional, Dict, Any # <-- ENSURE Dict and Any are imported
# In routers/timesheet.py

# In routers/timesheet.py

@router.patch("/{timesheet_id}/review", response_model=schemas.Timesheet)
def update_timesheet_review(
    timesheet_id: int,
    payload: Dict[str, Any],
    db: Session = Depends(get_db)
):
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    if not ts:
        raise HTTPException(status_code=404, detail="Timesheet not found")

    # 1. Get the current data dictionary
    current_data = ts.data if ts.data is not None else {}


    # Defensive check (if frontend accidentally sends string JSON)
    if isinstance(current_data, str):
        try:
            current_data = json.loads(current_data)
        except json.JSONDecodeError:
            current_data = {}

    # 2. Merge incoming payload
    current_data.update(payload)

    # Keep job name synced
    job_name = (
        current_data.get("job_name")
        or (current_data.get("job") or {}).get("job_description")
        or ts.timesheet_name
    )
    current_data["job_name"] = job_name

    # 3. Assign updated dictionary
    ts.data = current_data
    ts.timesheet_name = job_name

    # Flag as modified (required for SQLAlchemy JSONB)
    flag_modified(ts, "data")

    # 4. Commit and Refresh
    db.commit()
    db.refresh(ts)

    return ts

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os


class NotificationRequest(BaseModel):
    email: str
    subject: str
    message: str

# -------------------------------
# Send Notification Endpoint
# -------------------------------
@router.post("/send-notification")
def send_notification(request: NotificationRequest):
    if not sender_email or not sender_password:
        raise HTTPException(status_code=500, detail="SMTP credentials not set or invalid.")

    try:
        # Create email
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = request.email
        msg['Subject'] = request.subject
        msg.attach(MIMEText(request.message, 'plain'))

        # Send via Gmail SMTP
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()

        return {"status": "sent", "to": request.email}

    except smtplib.SMTPAuthenticationError as auth_err:
        raise HTTPException(
            status_code=500,
            detail=f"SMTP Authentication failed: {auth_err}"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")
    



@router.post("/{timesheet_id}/resend/", response_model=dict)
def resend_timesheet(timesheet_id: int, db: Session = Depends(get_db)):
    """
    Resend a timesheet (e.g., re-process, notify stakeholders, etc.)
    """
    try:
        # Fetch the timesheet from the database
        timesheet = db.query(Timesheet).filter(Timesheet.id == timesheet_id).first()
        if not timesheet:
            raise HTTPException(status_code=404, detail="Timesheet not found")

        # Convert to dict safely using as_dict
        data = timesheet.as_dict()

        # Extract safe fields
        foreman_name = data.get("foreman_name") or "Unknown Foreman"
        job_code = data.get("job_code") or "N/A"
        phase_name = data.get("phase_name") or "N/A"

        # Log the resend action
        print(f"🔄 RESENDING TIMESHEET ID {timesheet_id} - Foreman: {foreman_name}, Job: {job_code}, Phase: {phase_name}")

        # TODO: Add your actual resend logic here (email, notifications, OCR re-processing, etc.)
        # Example:
        # send_email(timesheet)
        # process_ocr_again(timesheet)
        # notify_foreman(timesheet)

        # Optionally update a "last_resent_at" timestamp
        # timesheet.last_resent_at = datetime.utcnow()
        # db.commit()

        return {
            "message": "Timesheet resent successfully",
            "timesheet_id": timesheet_id,
            "foreman": foreman_name,
            "job_code": job_code,
            "phase_name": phase_name
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Resend error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to resend timesheet")
    

import os
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ---------------- UTILS ----------------

# def write_safe(ws, row, col, value):
#     """
#     Writes a value to a cell, handling merged cells correctly.
#     Always writes to the top-left of a merged range.
#     """
#     cell = ws.cell(row=row, column=col)
#     for merged_range in ws.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             top_left = merged_range.start_cell
#             ws.cell(row=top_left.row, column=top_left.col_idx).value = value
#             return
#     cell.value = value

# def clear_range(ws, start_row, end_row, start_col, end_col):
#     """
#     Clears values in a range but keeps formatting and merged cells.
#     """
#     for r in range(start_row, end_row + 1):
#         for c in range(start_col, end_col + 1):
#             cell = ws.cell(row=r, column=c)
#             if isinstance(cell, MergedCell):
#                 # Clear only top-left of merged cell
#                 for merged_range in ws.merged_cells.ranges:
#                     if cell.coordinate in merged_range:
#                         top_left = merged_range.start_cell
#                         ws.cell(row=top_left.row, column=top_left.col_idx).value = None
#                         break
#             else:
#                 cell.value = None

# # ---------------- ENDPOINT ----------------


# @router.patch("/{timesheet_id}/pe-review")
# def pe_review_and_generate_excel(
#     timesheet_id: int,
#     payload: dict,
#     request: Request,
#     db: Session = Depends(get_db),
#     current_user: models.User = Depends(oauth2.get_current_user)
# ):
#     print("\n========== PE REVIEW START ==========")

#     # 1. Fetch and Update DB 
#     ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
#     if not ts:
#         raise HTTPException(status_code=404, detail="Timesheet not found")

#     ts.data.update(payload)
#     ts.status = "APPROVED_BY_PE"
#     flag_modified(ts, "data")
#     db.commit()

#     # 2. File Path Setup
#     BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#     template_path = os.path.join(BASE_DIR, "assets", "Alfonso - Template.xlsx")
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     file_name = f"Final_TS_{ts.id}_{timestamp}.xlsx"
#     output_path = os.path.join(BASE_DIR, "storage", "pe_reports", file_name)
#     os.makedirs(os.path.dirname(output_path), exist_ok=True)

#     # 3. Load Workbook
#     wb = load_workbook(template_path)
#     ws = wb["Timesheet"]

#     # 4. Header Information
#     write_safe(ws, 1, 4, str(ts.date))                       # Date (D1)
#     write_safe(ws, 3, 21, ts.data.get("job_name", ""))       # Job Name (U3)
#     write_safe(ws, 5, 21, ts.data.get("location", ""))       # Location (U5)
#     write_safe(ws, 5, 17, ts.data.get("project_engineer", "")) # PE Name (Q5)
#     write_safe(ws, 6, 17, current_user.username)             # Approved By (Q6)

#     # 5. Phase Code Mapping 
#     # Use selectedPhases from data to ensure columns match what was entered
#     phase_codes = ts.data.get("selectedPhases", [])
#     phase_col_map = {}
#     start_col = 8  # Column H in Alfonso template for Phase headers
#     for idx, phase in enumerate(phase_codes[:10]): # Limit to 10 phases as per template
#         col = start_col + (idx * 2) # Phases are usually spaced (REG/SB groups)
#         phase_col_map[str(phase)] = col
#         write_safe(ws, 8, col, str(phase))

#     # 6. Write Employees (Starts Row 11, step of 2) 
#     row_idx = 11
#     for emp in ts.data.get("employees", []):
#         write_safe(ws, row_idx, 1, emp.get("id")) # EMP #
#         name = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
#         write_safe(ws, row_idx, 2, name) # Name
        
#         # Handle Nested Hours: {"phase": {"class": hours}}
#         hours_per_phase = emp.get("hours_per_phase", {})
#         for phase, classes in hours_per_phase.items():
#             col = phase_col_map.get(str(phase))
#             if col and isinstance(classes, dict):
#                 total_hours = sum(float(h) for h in classes.values())
#                 write_safe(ws, row_idx, col, total_hours)
#         row_idx += 2 # Alfonso template has a sub-row for classes, move to next employee slot

#     # 7. Write Equipment (Starts Row 35) 
#     eq_row = 35
#     for eq in ts.data.get("equipment", []):
#         write_safe(ws, eq_row, 1, eq.get("name"))
#         write_safe(ws, eq_row, 2, eq.get("id"))
        
#         eq_hours = eq.get("hours_per_phase", {})
#         for phase, vals in eq_hours.items():
#             col = phase_col_map.get(str(phase))
#             if col and isinstance(vals, dict):
#                 write_safe(ws, eq_row, col, vals.get("REG", 0)) # REG column
#                 write_safe(ws, eq_row, col + 1, vals.get("SB", 0)) # SB column
#         eq_row += 1

#     # 8. Vendors and Materials (Row 51+) 
#     v_row = 51
#     for v in ts.data.get("vendors", []):
#         v_name = f"{v.get('vendor_name')} - {v.get('material_name')}"
#         write_safe(ws, v_row, 1, v_name)
        
#         # Extract tickets from {"ID": value} dict
#         tickets = v.get("tickets_loads", 0)
#         if isinstance(tickets, dict):
#             tickets = list(tickets.values())[0] if tickets else 0
#         write_safe(ws, v_row, 4, tickets)
        
#         # Quantities per phase
#         v_qty = v.get("hours_per_phase", {})
#         for phase, qty in v_qty.items():
#             col = phase_col_map.get(str(phase))
#             if col:
#                 write_safe(ws, v_row, col, float(qty))
#         v_row += 1

#     # 9. Save and Record
#     wb.save(output_path)
#     base_url = str(request.base_url).rstrip("/")
#     file_url = f"{base_url}/storage/pe_reports/{file_name}"
#     db.add(models.TimesheetFile(timesheet_id=ts.id, file_path=file_url))
#     db.commit()

#     print("✅ EXCEL SAVED:", output_path)
#     return ts


def write_safe(ws, row, col, value):
    """
    Writes a value to a cell, handling merged cells correctly.
    Always writes to the top-left of a merged range.
    """
    if value is None:
        value = ""
        
    cell = ws.cell(row=row, column=col)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Get top-left cell of the merged range
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left_cell.value = value
            return
    cell.value = value

# ---------------- ENDPOINT ----------------

@router.patch("/{timesheet_id}/pe-review")
def pe_review_and_generate_excel(
    timesheet_id: int,
    payload: dict,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(oauth2.get_current_user)
):
    print("\n========== PE REVIEW START ==========")

    # 1. Fetch and Update DB 
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    if not ts:
        raise HTTPException(status_code=404, detail="Timesheet not found")

    # Update the data field with the incoming payload (which matches your JSON structure)
    ts.data.update(payload)
    ts.status = "APPROVED_BY_PE"
    flag_modified(ts, "data")
    db.commit()

    # 2. File Path Setup
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(BASE_DIR, "assets", "Alfonso - Template.xlsx")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"Final_TS_{ts.id}_{timestamp}.xlsx"
    output_path = os.path.join(BASE_DIR, "storage", "pe_reports", file_name)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 3. Load Workbook
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Excel template not found")
        
    wb = load_workbook(template_path)
    ws = wb["Timesheet"]

    # 4. Header Information
    write_safe(ws, 1, 4, str(ts.date))                        # Date (D1)
    write_safe(ws, 3, 21, ts.data.get("job_name", ""))        # Job Name (U3)
    write_safe(ws, 5, 21, ts.data.get("location", ""))        # Location (U5)
    write_safe(ws, 5, 17, ts.data.get("project_engineer", "")) # PE Name (Q5)
    write_safe(ws, 6, 17, current_user.username)              # Approved By (Q6)

    # 5. Phase Code Mapping (Selected Phases)
    # Mapping codes like "110350" to specific columns in the template
    phase_codes = ts.data.get("selectedPhases", [])
    phase_col_map = {}
    start_col = 8  # Column H 
    
    for idx, phase in enumerate(phase_codes[:10]): 
        col = start_col + (idx * 2) 
        p_code = str(phase).strip()
        phase_col_map[p_code] = col
        write_safe(ws, 8, col, p_code)

    # 6. Write Employees (Row 11, step of 2)
    row_idx = 11
    for emp in ts.data.get("employees", []):
        write_safe(ws, row_idx, 1, emp.get("id")) 
        full_name = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
        write_safe(ws, row_idx, 2, full_name) 
        
        # Parse nested hours: {"phase_code": {"class_code": hours}}
        emp_hours = emp.get("hours_per_phase", {})
        for p_code, class_data in emp_hours.items():
            col = phase_col_map.get(str(p_code))
            if col:
                # Sum hours across all labor classes for this phase
                if isinstance(class_data, dict):
                    total_h = sum(float(h or 0) for h in class_data.values())
                else:
                    total_h = float(class_data or 0)
                write_safe(ws, row_idx, col, total_h)
        row_idx += 2 

    # 7. Write Equipment (Row 35)
    eq_row = 35
    for eq in ts.data.get("equipment", []):
        write_safe(ws, eq_row, 1, eq.get("name"))
        write_safe(ws, eq_row, 2, eq.get("id"))
        
        eq_hours = eq.get("hours_per_phase", {})
        for p_code, vals in eq_hours.items():
            col = phase_col_map.get(str(p_code))
            if col and isinstance(vals, dict):
                write_safe(ws, eq_row, col, float(vals.get("REG", 0)))
                write_safe(ws, eq_row, col + 1, float(vals.get("SB", 0)))
        eq_row += 1

    # 8. Vendors and Materials (Row 51+)
    v_row = 51
    for v in ts.data.get("vendors", []):
        v_display = f"{v.get('vendor_name', '')} - {v.get('material_name', '')}"
        write_safe(ws, v_row, 1, v_display)
        
        # Handle tickets_loads: {"91_2": 2}
        tickets_data = v.get("tickets_loads", 0)
        if isinstance(tickets_data, dict):
            # Extract first value from the dict
            tickets_val = next(iter(tickets_data.values())) if tickets_data else 0
        else:
            tickets_val = tickets_data
        write_safe(ws, v_row, 4, tickets_val)
        
        # Map quantities per phase
        v_qty_per_phase = v.get("hours_per_phase", {})
        for p_code, qty in v_qty_per_phase.items():
            col = phase_col_map.get(str(p_code))
            if col:
                write_safe(ws, v_row, col, float(qty or 0))
        v_row += 1

    # 9. Save and Record
    wb.save(output_path)
    base_url = str(request.base_url).rstrip("/")
    file_url = f"{base_url}/storage/pe_reports/{file_name}"
    
    new_file_record = models.TimesheetFile(timesheet_id=ts.id, file_path=file_url)
    db.add(new_file_record)
    db.commit()

    print("✅ EXCEL GENERATED AND SAVED:", output_path)
    return {"status": "success", "file_url": file_url, "timesheet_id": ts.id}


@router.post("/{timesheet_id}/notify-executive")
def notify_natalia(timesheet_id: int, db: Session = Depends(get_db)):
    ts = db.query(models.Timesheet).filter(models.Timesheet.id == timesheet_id).first()
    foreman = db.query(models.User).filter(models.User.id == ts.foreman_id).first()
    
    foreman_name = f"{foreman.first_name}"
    alert_msg = format_natalia_alert(ts.data, foreman_name)
    
    # Logic to send push notification or update Natalia's Dashboard
    # For now, we reuse your existing email notification
    notification = NotificationRequest(
        email="natalia@mluisconstruction.com", 
        subject="📍 Site Activity Alert",
        message=alert_msg
    )
    return send_notification(notification)
